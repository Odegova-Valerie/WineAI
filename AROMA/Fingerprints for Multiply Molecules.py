import pandas as pd
import numpy as np
import seaborn as sn
from rdkit import Chem
from rdkit.Chem import AllChem
from tqdm import tqdm
from rdkit.Chem import rdMolDescriptors
import openpyxl

file_path=r"C:\Users\Госпожа Виктория\Desktop\basic DATASET.xlsx"
sheet_name = 'Хим состав для Multiple'
dataset = pd.read_excel(file_path, sheet_name = sheet_name)
dataset = dataset.reset_index(drop=True)
def combine_values(row):
    result='.'.join(str(val) for val in row if val!=0)

    return result

dataset['Combined']=dataset.apply(combine_values, axis=1)
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    dataset.to_excel(writer, sheet_name='Хим состав для Multiple', index=False)

def find_empty_column(sheet, start_column):
    for column_index in range(start_column, sheet.max_column + 2):
        if all([cell.value is None for cell in sheet[f'{openpyxl.utils.get_column_letter(column_index)}'][2:]]):
            return column_index
    return sheet.max_column + 1

def generate_fpts(dataset, file_path, start_column=51):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Хим состав для Multiple']
    for row_index, smile in enumerate(dataset['Combined'].values, start=2):
        mol = Chem.MolFromSmiles(smile)
        Mfpts = []
        #print(smile)
        if mol is not None:
            mfpt = AllChem.GetMorganFingerprintAsBitVect(mol, 2, nBits=2048)
            Mfpts.append(mfpt.ToBitString())
            Mfpts_row = [int(x) for x in Mfpts]
            column_index = find_empty_column(sheet, start_column)
            for index, value in enumerate(row_data, start=2):
                for column_index, x in enumerate(Mfpts, start=51):
                    sheet.cell(row=row_index, column=column_index+index-1, value=x)
        else:
            print(f"Ошибка преобразования   {smile}")
    workbook.save(file_path)       
    return file_path
#morgan_fpts = generate_fpts(dataset)

#dataset['Combined Multiple']=dataset.apply(generate_fpts, axis=1)
#print(morgan_fpts[0])
tmp_lst=[]
for num in morgan_fpts[0]:
    tmp_lst.append(int(num))
#print(tmp_lst)'''
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    dataset.to_excel(writer, sheet_name='Хим состав для Multiple', index=False)
